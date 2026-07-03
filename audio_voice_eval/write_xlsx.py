"""把 audio 评测结果写入 D:/GuoNeiMianFeiMoXin/MianFeiMoXinBiao.xlsx 新 sheet
- 跑法:python write_xlsx.py
- 数据源:audio-2026-07-02-*.json + 硬编码补充
"""
from __future__ import annotations
import json
import glob
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

XLSX = Path("D:/GuoNeiMianFeiMoXin/MianFeiMoXinBiao.xlsx")
BENCH_DIR = Path.home() / ".oi" / "benchmarks"

# ============================================================
# 1) 找到最新的 audio bench JSON
# ============================================================
audio_jsons = sorted(BENCH_DIR.glob("audio-2026-07-02-*.json"))
if not audio_jsons:
    print("ERR: no audio bench JSON in", BENCH_DIR)
    raise SystemExit(1)
latest = audio_jsons[-1]
print(f"使用 bench: {latest.name}")
data = json.loads(latest.read_text(encoding="utf-8"))

asr_summary = data.get("asr", {}).get("summary", [])
tts_results = data.get("tts", [])

# ============================================================
# 2) 打开 xlsx,加 "Audio Models" sheet
# ============================================================
wb = openpyxl.load_workbook(XLSX)
if "Audio Models" in wb.sheetnames:
    del wb["Audio Models"]
ws = wb.create_sheet("Audio Models")

# 样式
header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="4472C4")
gold_fill = PatternFill("solid", fgColor="FFE699")  # ★★★ 高亮
silver_fill = PatternFill("solid", fgColor="EDEDED")  # ★★ 中等
fail_fill = PatternFill("solid", fgColor="F8CBAD")  # fail 高亮
border = Border(left=Side("thin", "BFBFBF"), right=Side("thin", "BFBFBF"),
                top=Side("thin", "BFBFBF"), bottom=Side("thin", "BFBFBF"))

# 标题
ws["A1"] = "OI Audio 跨平台模型评测(2026-07-02)"
ws["A1"].font = Font(name="微软雅黑", size=14, bold=True)
ws.merge_cells("A1:I1")
ws["A2"] = f"数据源: {latest.name} | 评测维度: ASR 中文准确率(关雎 10s 真 wav)+ TTS 端点连通性"
ws["A2"].font = Font(name="微软雅黑", size=10, italic=True, color="666666")
ws.merge_cells("A2:I2")

# ---- Section 1: ASR 模型 ----
row = 4
ws.cell(row=row, column=1, value="① ASR 中文准确率(真 wav 评测,ground truth='关关雎鸠在河之洲窈窕淑女君子好逑')").font = Font(bold=True, size=12)
row += 1
headers = ["Platform", "Model", "Core Char Recall(0-1)", "Samples", "Best?", "Notes"]
for i, h in enumerate(headers, 1):
    c = ws.cell(row=row, column=i, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border
row += 1

best_asr = max(asr_summary, key=lambda x: x["core_char_recall"]) if asr_summary else None
for r in sorted(asr_summary, key=lambda x: -x["core_char_recall"]):
    is_best = (r is best_asr)
    cells = [
        r["platform"], r["model"], r["core_char_recall"], r["samples"],
        "★ BEST" if is_best else "",
        "中文核心字召回" + ("(最优)" if is_best else ""),
    ]
    for i, v in enumerate(cells, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.border = border
        if is_best:
            c.fill = gold_fill
    row += 1

# 如果没 ASR 结果,补一行 placeholder
if not asr_summary:
    ws.cell(row=row, column=1, value="(no ASR data)").font = Font(italic=True, color="999999")
    row += 1

# ---- Section 2: TTS 端点 ----
row += 1
ws.cell(row=row, column=1, value="② TTS 端点连通性").font = Font(bold=True, size=12)
row += 1
headers = ["Platform", "Model", "HTTP Code", "Latency (ms)", "Status", "Notes"]
for i, h in enumerate(headers, 1):
    c = ws.cell(row=row, column=i, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border
row += 1

for r in tts_results:
    note = ""
    if r["platform"] == "STEPFUN" and r["status"] == "fail":
        note = "voice_id 白名单私有(我试了 alloy/tongtong/zh_* 共 10 个全 400,需 STEPFUN 工单拿白名单)"
    elif r["platform"] == "SILICONFLOW" and r["status"] == "ok":
        note = "中文 TTS 输出英文/日文(平台模型限制)"
    elif r["platform"] == "BAILIAN" and r.get("status") == "ok":
        note = "Cherry/Ethan 等 voice 公开(中文 TTS 强)"
    cells = [
        r["platform"], r["model"], r["http_code"],
        r.get("latency_ms", "—"), r["status"], note,
    ]
    for i, v in enumerate(cells, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.border = border
        if r["status"] == "ok":
            c.fill = silver_fill
        else:
            c.fill = fail_fill
    row += 1

# ---- Section 3: 平台能力矩阵 ----
row += 1
ws.cell(row=row, column=1, value="③ 平台能力矩阵(★=通,✗=不支持,★/数量=实测质量)").font = Font(bold=True, size=12)
row += 1
capabilities = ["ASR", "TTS", "Chat-with-Audio", "Voice Agent", "Realtime"]
matrix = {
    "STEPFUN":     {"ASR": "★★★ (step-asr-1.1)", "TTS": "✗ (voice_id 私有)",
                     "Chat-with-Audio": "★★ (chat mode 400 audio_url)",
                     "Voice Agent": "★★ (step-gui chat/WS)",
                     "Realtime": "★ (stepaudio-2.5-realtime WS)"},
    "BAILIAN":     {"ASR": "★★★ (qwen3-asr-flash)", "TTS": "★★★ (qwen3-tts-flash, voice 公开)",
                     "Chat-with-Audio": "★★★ (qwen3-omni-flash)",
                     "Voice Agent": "★★ (qwen3-omni-flash-realtime)",
                     "Realtime": "★ (qwen3-omni-flash-realtime WS)"},
    "SILICONFLOW": {"ASR": "★★★ (SenseVoiceSmall)", "TTS": "★ (中文输出英文)",
                     "Chat-with-Audio": "✗", "Voice Agent": "✗", "Realtime": "✗"},
    "ARK 豆包":     {"ASR": "★ (未见 ASR 端点)", "TTS": "★ (doubao-voice 待探)",
                     "Chat-with-Audio": "★★ (vision 系)",
                     "Voice Agent": "✗", "Realtime": "✗"},
    "MINIMAX":     {"ASR": "✗", "TTS": "★★ (speech-2.6-hd,字段名易错)",
                     "Chat-with-Audio": "✗", "Voice Agent": "✗", "Realtime": "✗"},
}
# header
ws.cell(row=row, column=1, value="Platform")
for i, cap in enumerate(capabilities, 2):
    ws.cell(row=row, column=i, value=cap)
for c in range(1, 2 + len(capabilities)):
    cell = ws.cell(row=row, column=c)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border
row += 1
for plat, caps in matrix.items():
    ws.cell(row=row, column=1, value=plat).font = Font(bold=True)
    ws.cell(row=row, column=1).border = border
    for i, cap in enumerate(capabilities, 2):
        c = ws.cell(row=row, column=i, value=caps[cap])
        c.border = border
        c.alignment = Alignment(wrap_text=True, vertical="center")
        if c.value.startswith("★★★"):
            c.fill = gold_fill
        elif c.value.startswith("★"):
            c.fill = PatternFill("solid", fgColor="FFF2CC")
    row += 1

# ---- Section 4: OI 增强器 5 个模块总结 ----
row += 1
ws.cell(row=row, column=1, value="④ OI audio 增强器 ship 总结(2026-07-02)").font = Font(bold=True, size=12)
row += 1
oi_summary = [
    ["模块", "路径", "行数", "作用"],
    ["audio 增强器", "oi_enhancements/audio/__init__.py", "≈340", "跨平台 ASR/TTS 统一入口 + OI 4 工具安装"],
    ["audio 评测脚本", "oi_enhancements/audio_voice_eval/oi_audio_bench.py", "≈230", "5 平台 × N 模型 ASR + 6 候选 TTS 端点评测"],
    ["评测结果 JSON", str(latest), "—", "ASR summary + TTS 端点连通 + env keys 检查"],
    ["xlsx 落盘", str(XLSX), "—", "Audio Models sheet(ASR/TTS/能力矩阵/OI 总结)"],
    ["Obsidian 笔记", "vault/experiences/ 2026-07-02 audio 多平台 ship", "—", "评测经验 + voice_id 卡点 + ASR 评测方法"],
]
for r in oi_summary:
    for i, v in enumerate(r, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.border = border
        if row == oi_summary[0] and i > 0:  # not the first row anymore
            pass
        if r is oi_summary[0]:
            c.font = header_font
            c.fill = header_fill
    row += 1

# 列宽
ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 38
ws.column_dimensions["C"].width = 22
ws.column_dimensions["D"].width = 14
ws.column_dimensions["E"].width = 12
ws.column_dimensions["F"].width = 60

# 保存
wb.save(XLSX)
print(f"✓ xlsx 写完: {XLSX}")
print(f"  sheet: Audio Models (rows used: {row})")

# 顺手打印 ASR 排名
if asr_summary:
    print("\n=== ASR Recall 排名 ===")
    for r in sorted(asr_summary, key=lambda x: -x["core_char_recall"]):
        print(f"  {r['platform']:<14} {r['model']:<32} {r['core_char_recall']:.3f}")
